# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
LingXing (领星) portal automation module.

This module provides automation utilities for the LingXing ERP system,
including login, navigation, and common operations.

Usage:
    from domains.lingxing import LingXingPortal

    portal = LingXingPortal(page)
    portal.login(username, password)
    portal.navigate_to_shipment(order_no)
"""

from __future__ import annotations

from typing import Optional

from playwright.sync_api import Page

from core.browser.interaction import (
    human_pause,
    click_first_visible,
    fill_first,
    wait_for_stability,
    wait_for_any_selector,
)
from ..selectors import (
    BASIC_SELECTORS,
    FBA_DELIVERY_ORDER_SELECTORS,
    LOGISTICS_ADD_CUSTOMS_CLEARANCE_SELECTORS,
    LOGISTICS_CUSTOMS_MANAGEMENT_SELECTORS,
)

LINGXING_SELECTORS = dict(BASIC_SELECTORS)
LINGXING_SELECTORS.update(FBA_DELIVERY_ORDER_SELECTORS)
LINGXING_SELECTORS.update(LOGISTICS_CUSTOMS_MANAGEMENT_SELECTORS)
LINGXING_SELECTORS.update(LOGISTICS_ADD_CUSTOMS_CLEARANCE_SELECTORS)
from ... import config


class LingXingPortal:
    """Automation wrapper for LingXing ERP system."""

    def __init__(self, page: Page, base_url: str = None):
        """
        Initialize the LingXing portal wrapper.

        Args:
            page: Playwright Page object.
            base_url: Base URL for the LingXing ERP system. Defaults to config.LINGXING_URL.
        """
        self.page = page
        self.base_url = base_url or config.LINGXING_URL

    def goto_home(self) -> None:
        """Navigate to LingXing home page."""
        import logging
        logger = logging.getLogger("LingXingPortal")
        logger.info(f"正在跳转到领星首页: {self.base_url}")
        self.page.goto(self.base_url, wait_until="domcontentloaded", timeout=30000)
        wait_for_stability(self.page)

    def is_login_page(self) -> bool:
        """Check if current page is login page."""
        selectors = LINGXING_SELECTORS.get("account_input", [])
        for selector in selectors:
            try:
                locator = self.page.locator(selector)
                if locator.count() > 0:
                    return True
            except Exception:
                continue
        return False

    def login(self, username: str, password: str) -> bool:
        """
        Perform login to LingXing.

        Args:
            username: LingXing account username/email.
            password: Account password.

        Returns:
            True if login was attempted, False if not on login page.
        """
        if not self.is_login_page():
            return False

        human_pause(500, 1200)
        fill_first(
            self.page,
            LINGXING_SELECTORS.get("account_input", []),
            username,
        )
        human_pause(300, 800)
        fill_first(
            self.page,
            LINGXING_SELECTORS.get("password_input", []),
            password,
        )
        human_pause(500, 1200)
        click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("login_button", []),
        )
        wait_for_stability(self.page, timeout_ms=15000)
        return True

    def ensure_language_cn(self) -> bool:
        """Switch language to Simplified Chinese if needed."""
        # TODO: Implement language switch logic
        return True

    def navigate_to_shipment(self, order_no: str) -> None:
        """
        Navigate to shipment detail page.
        
        Args:
            order_no: Shipment order number (e.g., SP260119001)
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        url = config.LINGXING_DELIVERY_URL.format(order_no=order_no)
        logger.info(f"正在尝试跳转到发货单详情页: {url}")
        self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
        wait_for_stability(self.page)
        human_pause(500, 1000)
        logger.info(f"跳转完成，当前页面标题: {self.page.title()}")

    def navigate_to_shipment_edit(self, order_no: str) -> None:
        """
        Navigate to shipment edit page.

        Args:
            order_no: Shipment order number (e.g., SP260119001)
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        url = config.LINGXING_DELIVERY_EDIT_URL.format(order_no=order_no)
        logger.info(f"正在尝试跳转到发货单编辑页: {url}")
        self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
        wait_for_stability(self.page)
        human_pause(500, 1000)
        # Wait for key elements on the edit page to appear before proceeding
        ready_selectors = []
        ready_selectors.extend(LINGXING_SELECTORS.get("packing_info_tab", []))
        ready_selectors.extend(LINGXING_SELECTORS.get("logistics_info_tab", []))
        if ready_selectors:
            if not wait_for_any_selector(self.page, ready_selectors, timeout_ms=10000):
                logger.warning(f"{order_no}: 发货单编辑页关键元素未出现，页面可能未完全加载")
        logger.info(f"跳转完成，当前页面标题: {self.page.title()}")

    def click_edit(self) -> bool:
        """Click the edit button on shipment detail page."""
        import logging
        logger = logging.getLogger("LingXingPortal")
        logger.info("准备点击编辑按钮...")
        success = click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("edit_button", []),
        )
        if success:
            logger.info("成功点击编辑按钮")
        else:
            logger.error("未能找到或点击编辑按钮")
        return success

    def click_packing_info_tab(self) -> bool:
        """Click the packing info tab."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("packing_info_tab", []),
        )

    def click_get_shipment_packing(self) -> bool:
        """Click the 'Get Shipment Packing' button."""
        human_pause(500, 800)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("get_shipment_packing_btn", []),
        )

    def confirm_get_shipment_packing(self) -> bool:
        """Click confirm button after getting shipment packing."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("get_shipment_packing_confirm_btn", []),
        )

    def get_shipment_quantity(self) -> tuple[str, str] | None:
        """
        Get shipment quantity text and parse it.
        
        Returns:
            Tuple of (outer_value, inner_value) e.g., ("945", "945") from "945 (945)"
            or None if not found.
        """
        import re
        selectors = LINGXING_SELECTORS.get("shipment_quantity", [])
        for selector in selectors:
            try:
                locator = self.page.locator(selector)
                if locator.count() > 0 and locator.first.is_visible():
                    text = locator.first.text_content()
                    if text:
                        # Parse format: "945 (945)"
                        match = re.match(r"(\d+)\s*\((\d+)\)", text.strip())
                        if match:
                            return (match.group(1), match.group(2))
            except Exception:
                continue
        return None

    def check_quantity_match(self) -> tuple[bool, str]:
        """
        Check if shipment quantity matches.
        
        Returns:
            Tuple of (is_match, message)
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        
        result = self.get_shipment_quantity()
        if result is None:
            logger.warning("未能从页面获取到发货量数据元素")
            return (False, "无法获取发货量数据")
        
        outer, inner = result
        logger.info(f"抓取到装箱数量数据: 计划发货量={outer}, 装箱总数={inner}")
        
        if outer == inner:
            return (True, f"发货量一致: {outer}")
        else:
            return (False, f"发货量不一致: {outer} vs {inner}")

    def click_logistics_info_tab(self) -> bool:
        """Click the logistics info tab."""
        import logging
        logger = logging.getLogger("LingXingPortal")
        human_pause(300, 600)
        success = click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("logistics_info_tab", []),
        )
        if success:
            logger.info("成功点击物流信息页签")
        else:
            logger.error("未能找到并点击物流信息页签")
        return success

    def fill_logistics_numbers(self, logistics_no: str, tracking_no: str) -> bool:
        """
        Fill in logistics number and tracking number.
        
        Args:
            logistics_no: Logistics provider number.
            tracking_no: Tracking number.
        
        Returns:
            True if both fields were filled successfully.
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        
        human_pause(300, 600)
        logger.info(f"正在填写物流商单号: {logistics_no}")
        filled_logistics = fill_first(
            self.page,
            LINGXING_SELECTORS.get("logistics_number_input", []),
            logistics_no,
        )
        if not filled_logistics:
            logger.error("填写物流商单号失败")
        
        human_pause(300, 600)
        logger.info(f"正在填写查询单号: {tracking_no}")
        filled_tracking = fill_first(
            self.page,
            LINGXING_SELECTORS.get("tracking_number_input", []),
            tracking_no,
        )
        if not filled_tracking:
            try:
                panel = self.page.locator('//*[@id="pane-物流信息"]')
                base = panel.first if panel.count() > 0 else self.page
                inputs = base.locator(
                    '//tr[contains(@class,"vxe-body--row")]//td[2]//input[@maxlength="255"]'
                )
                count = inputs.count()
                logger.info(f"查询单号备用定位: 找到 {count} 个 行内第二列 输入框")
                target = inputs.nth(0) if count > 0 else None
                if target and target.is_visible():
                    try:
                        target.click()
                        self.page.keyboard.press("Control+A")
                        self.page.keyboard.press("Backspace")
                    except Exception:
                        pass
                    try:
                        target.press_sequential(tracking_no, delay=100)
                    except Exception:
                        target.fill(tracking_no)
                    filled_tracking = True
            except Exception:
                filled_tracking = False
            logger.error("填写查询单号失败")
            
        return filled_logistics and filled_tracking

    def confirm_logistics_info(self) -> bool:
        """Click confirm button to save logistics info."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("confirm_button", []),
        )

    def _normalize_text(self, value: Optional[str]) -> str:
        if not value:
            return ""
        return " ".join(value.split())

    def _run_async(self, coro):
        import asyncio
        from concurrent.futures import ThreadPoolExecutor

        try:
            asyncio.get_running_loop()
        except RuntimeError:
            return asyncio.run(coro)
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(asyncio.run, coro)
            return future.result()

    def _looks_like_id(self, value: str) -> bool:
        return value.isdigit() and len(value) >= 6

    def _xpath_literal(self, value: str) -> str:
        if "'" not in value:
            return f"'{value}'"
        if '"' not in value:
            return f'"{value}"'
        parts = value.split("'")
        return "concat(" + ", \"'\", ".join([f"'{p}'" for p in parts]) + ")"

    def _get_select_display_value(self, selectors: list[str]) -> Optional[str]:
        """
        Get displayed text for a readonly select field.
        """
        for selector in selectors:
            try:
                locator = self.page.locator(selector)
                if locator.count() > 0:
                    element = locator.first
                    text = self._normalize_text(element.text_content())
                    if text and not self._looks_like_id(text):
                        return text
                    # If we landed on the input itself, try to fetch the visible select label.
                    ancestor_input = element.locator(
                        "xpath=ancestor::div[contains(@class,'el-input')]"
                    )
                    if ancestor_input.count() > 0:
                        label = ancestor_input.first.locator(
                            ".fake-placeholder-hidden, .fake-select-label"
                        )
                        if label.count() > 0:
                            label_text = self._normalize_text(label.first.text_content())
                            if label_text:
                                return label_text
                    value = self._normalize_text(element.get_attribute("value"))
                    if value and not self._looks_like_id(value):
                        return value
            except Exception:
                continue
        return None

    def _wait_for_select_display(self, selector_key: str, timeout_ms: int = 8000) -> Optional[str]:
        """
        Wait until a select shows a non-empty display text (not a numeric id).
        """
        import time

        end = time.monotonic() + timeout_ms / 1000.0
        last = None
        while time.monotonic() < end:
            value = self._get_select_display_value(LINGXING_SELECTORS.get(selector_key, []))
            if value and not self._looks_like_id(value):
                return value
            last = value
            human_pause(200, 400)
        if last and not self._looks_like_id(last):
            return last
        return None

    def _get_logistics_panel_select(self, index: int):
        panel = self.page.locator('//*[@id="pane-物流信息"]')
        if panel.count() == 0:
            return None
        selects = panel.first.locator(".el-select")
        if selects.count() <= index:
            return None
        return selects.nth(index)

    def _open_select_by_index(self, index: int) -> bool:
        select = self._get_logistics_panel_select(index)
        if select is None:
            return False
        target = select.locator(".el-input__inner")
        if target.count() == 0:
            target = select
        human_pause(600, 1200)
        try:
            target.first.click()
        except Exception:
            return False
        wait_for_any_selector(
            self.page,
            [
                '//div[contains(@class,"el-select-dropdown") and not(contains(@style,"display: none"))]'
            ],
            timeout_ms=3000,
        )
        human_pause(500, 900)
        return True

    def _open_select(self, selectors: list[str]) -> bool:
        if not selectors:
            return False
        # Slow down to mimic human interaction when opening dropdowns
        human_pause(600, 1200)
        opened = click_first_visible(self.page, selectors)
        if not opened:
            return False
        wait_for_any_selector(
            self.page,
            [
                '//div[contains(@class,"el-select-dropdown") and not(contains(@style,"display: none"))]'
            ],
            timeout_ms=3000,
        )
        human_pause(500, 900)
        return True

    def _select_option_by_text(self, text: str) -> bool:
        human_pause(600, 1200)
        literal = self._xpath_literal(text)
        selectors = [
            f'//div[contains(@class,"el-select-dropdown") and not(contains(@style,"display: none"))]'
            f'//li[contains(@class,"el-select-dropdown__item")][.//span[normalize-space()={literal}]]',
            f'//div[contains(@class,"el-select-dropdown") and not(contains(@style,"display: none"))]'
            f'//li[contains(@class,"el-select-dropdown__item")][normalize-space(.)={literal}]',
            f'//div[contains(@class,"el-select-dropdown") and not(contains(@style,"display: none"))]'
            f'//li[contains(@class,"el-select-dropdown__item")][contains(normalize-space(.),{literal})]',
        ]
        clicked = click_first_visible(self.page, selectors)
        human_pause(500, 900)
        return clicked

    def _get_select_values_from_logistics_panel(self) -> list[str]:
        """
        Fallback: collect select display values under the logistics info panel.
        """
        values: list[str] = []
        try:
            panel = self.page.locator('//*[@id="pane-物流信息"]')
            if panel.count() == 0:
                return values
            selects = panel.first.locator(".el-select")
            for i in range(selects.count()):
                sel = selects.nth(i)
                label = sel.locator(".fake-placeholder-hidden, .fake-select-label")
                if label.count() > 0:
                    text = self._normalize_text(label.first.text_content())
                    if text:
                        values.append(text)
                        continue
                inner = sel.locator(".el-input__inner")
                if inner.count() > 0:
                    val = self._normalize_text(inner.first.get_attribute("value"))
                    if val:
                        values.append(val)
        except Exception:
            return values
        return values

    def check_logistics_profile(
        self,
        expected_transport_mode: Optional[str] = None,
        expected_logistics_channel: Optional[str] = None,
        expected_logistics_provider: Optional[str] = None,
        *,
        auto_fix: bool = False,
    ) -> tuple[bool, str]:
        """
        Check transport mode, logistics channel, and provider before filling numbers.

        If auto_fix is True, try to select the expected value when mismatch is found.
        """
        import logging
        logger = logging.getLogger("LingXingPortal")

        issues = []
        logger.info(
            "物流信息校验开始: 期望 运输方式=%s, 物流渠道=%s, 物流商=%s",
            expected_transport_mode,
            expected_logistics_channel,
            expected_logistics_provider,
        )

        def _get_actual(expected: Optional[str], selector_key: str) -> Optional[str]:
            if expected is None:
                return None
            actual = self._get_select_display_value(LINGXING_SELECTORS.get(selector_key, []))
            if actual and self._looks_like_id(actual):
                return None
            return actual

        def _compare(label: str, expected: Optional[str], actual: Optional[str]) -> None:
            if expected is None:
                return
            actual_norm = self._normalize_text(actual)
            expected_norm = self._normalize_text(expected)
            if not actual_norm:
                issues.append(f"{label}=空(期望:{expected_norm})")
            elif actual_norm != expected_norm:
                issues.append(f"{label}={actual_norm}(期望:{expected_norm})")

        transport_mode = (
            self._wait_for_select_display("transport_mode_value")
            if expected_transport_mode is not None
            else _get_actual(expected_transport_mode, "transport_mode_value")
        )
        logistics_channel = (
            self._wait_for_select_display("logistics_channel_value")
            if expected_logistics_channel is not None
            else _get_actual(expected_logistics_channel, "logistics_channel_value")
        )
        logistics_provider = (
            self._wait_for_select_display("logistics_provider_value")
            if expected_logistics_provider is not None
            else _get_actual(expected_logistics_provider, "logistics_provider_value")
        )
        logger.info(
            "初次读取到物流信息: 运输方式=%s, 物流渠道=%s, 物流商=%s",
            transport_mode,
            logistics_channel,
            logistics_provider,
        )

        if (
            (expected_transport_mode and not transport_mode)
            or (expected_logistics_channel and not logistics_channel)
            or (expected_logistics_provider and not logistics_provider)
        ):
            human_pause(600, 1200)
            transport_mode = transport_mode or _get_actual(expected_transport_mode, "transport_mode_value")
            logistics_channel = logistics_channel or _get_actual(expected_logistics_channel, "logistics_channel_value")
            logistics_provider = logistics_provider or _get_actual(expected_logistics_provider, "logistics_provider_value")
            logger.info(
                "二次读取到物流信息: 运输方式=%s, 物流渠道=%s, 物流商=%s",
                transport_mode,
                logistics_channel,
                logistics_provider,
            )

        if auto_fix:
            def _ensure(
                label: str,
                expected: Optional[str],
                actual: Optional[str],
                select_key: str,
                fallback_index: Optional[int],
            ) -> Optional[str]:
                if expected is None:
                    return actual
                if self._normalize_text(actual) == self._normalize_text(expected):
                    return actual
                logger.warning(f"{label} 当前值不匹配，尝试自动选择: {expected}")
                opened = self._open_select(LINGXING_SELECTORS.get(select_key, []))
                if not opened and fallback_index is not None:
                    opened = self._open_select_by_index(fallback_index)
                if not opened:
                    logger.error(f"{label} 选择框无法打开")
                    return actual
                if not self._select_option_by_text(expected):
                    logger.error(f"{label} 未找到选项: {expected}")
                    return actual
                wait_for_stability(self.page)
                new_value = self._get_select_display_value(
                    LINGXING_SELECTORS.get(select_key.replace("_select", "_value"), [])
                )
                if not new_value and fallback_index is not None:
                    fallback_values = self._get_select_values_from_logistics_panel()
                    if len(fallback_values) > fallback_index:
                        new_value = fallback_values[fallback_index]
                return new_value or expected

            def _needs_fix(expected: Optional[str], actual: Optional[str]) -> bool:
                return expected is not None and self._normalize_text(actual) != self._normalize_text(expected)

            needs_fix = (
                _needs_fix(expected_transport_mode, transport_mode)
                or _needs_fix(expected_logistics_channel, logistics_channel)
                or _needs_fix(expected_logistics_provider, logistics_provider)
            )

            if needs_fix:
                transport_mode = _ensure(
                    "运输方式",
                    expected_transport_mode,
                    transport_mode,
                    "transport_mode_select",
                    0,
                )
                logistics_channel = _ensure(
                    "物流渠道",
                    expected_logistics_channel,
                    logistics_channel,
                    "logistics_channel_select",
                    1,
                )
                logistics_provider = _ensure(
                    "物流商",
                    expected_logistics_provider,
                    logistics_provider,
                    "logistics_provider_select",
                    2,
                )
                logger.info(
                    "自动修正后物流信息: 运输方式=%s, 物流渠道=%s, 物流商=%s",
                    transport_mode,
                    logistics_channel,
                    logistics_provider,
                )

        # If any value is missing after auto-fix attempts, try a panel-order fallback.
        if (
            (expected_transport_mode and not transport_mode)
            or (expected_logistics_channel and not logistics_channel)
            or (expected_logistics_provider and not logistics_provider)
        ):
            fallback_values = self._get_select_values_from_logistics_panel()
            if len(fallback_values) >= 3:
                logger.info("未通过标签定位到物流信息，使用物流信息面板内选择框顺序回退")
                if expected_transport_mode and not transport_mode:
                    transport_mode = fallback_values[0]
                if expected_logistics_channel and not logistics_channel:
                    logistics_channel = fallback_values[1]
                if expected_logistics_provider and not logistics_provider:
                    logistics_provider = fallback_values[2]
                logger.info(
                    "回退读取到物流信息: 运输方式=%s, 物流渠道=%s, 物流商=%s",
                    transport_mode,
                    logistics_channel,
                    logistics_provider,
                )

        _compare("运输方式", expected_transport_mode, transport_mode)
        _compare("物流渠道", expected_logistics_channel, logistics_channel)
        _compare("物流商", expected_logistics_provider, logistics_provider)

        if issues:
            message = "物流信息校验失败: " + "; ".join(issues)
            logger.warning(message)
            return (False, message)

        logger.info(
            f"物流信息校验通过: 运输方式={transport_mode or '-'}, "
            f"物流渠道={logistics_channel or '-'}, 物流商={logistics_provider or '-'}"
        )
        return (True, "物流信息校验通过")

    def click_adjust(self) -> bool:
        """Click the adjust button."""
        human_pause(500, 1000)
        selectors = LINGXING_SELECTORS.get("adjust_button", [])
        if not selectors:
            return False
        # Wait for any adjust button selector to appear, not just the first one.
        if not wait_for_any_selector(self.page, selectors, timeout_ms=5000):
            return False
        return click_first_visible(self.page, selectors)

    def select_hangzhou_virtual_warehouse(self) -> bool:
        """Select '杭州虚拟仓' in shipment warehouse dropdown."""
        selectors = LINGXING_SELECTORS.get("shipment_warehouse_select", [])
        if not selectors:
            return False
        if not wait_for_any_selector(self.page, selectors, timeout_ms=5000):
            return False
        if not click_first_visible(self.page, selectors):
            return False
        human_pause(300, 600)
        option_selectors = LINGXING_SELECTORS.get("shipment_warehouse_option_hz_virtual", [])
        if not option_selectors:
            return False
        if not wait_for_any_selector(self.page, option_selectors, timeout_ms=5000):
            return False
        return click_first_visible(self.page, option_selectors)

    def has_empty_warehouse_info(self) -> bool:
        """
        Check if 'Shipment Warehouse Shop' or 'Shipment Warehouse FNSKU' columns have empty data.
        """
        import logging
        logger = logging.getLogger("LingXingPortal")

        if not self.wait_for_table_ready():
            logger.warning("表格可能未就绪，继续执行仓库信息检查")

        def first_nonempty_cells(selectors: list[str], label: str):
            first_locator = None
            first_selector = None
            for selector in selectors:
                try:
                    locator = self.page.locator(selector)
                    count = locator.count()
                    if count > 0 and first_locator is None:
                        first_locator = locator
                        first_selector = selector
                except Exception:
                    continue
            return first_locator, first_selector

        # Column: Shipment Warehouse Shop
        shop_cells, shop_selector = first_nonempty_cells(
            LINGXING_SELECTORS.get("warehouse_shop_cells", []),
            "店铺",
        )
        # Column: Shipment Warehouse FNSKU
        fnsku_cells, fnsku_selector = first_nonempty_cells(
            LINGXING_SELECTORS.get("warehouse_fnsku_cells", []),
            "FNSKU",
        )

        try:
            shop_count = shop_cells.count() if shop_cells else 0
            fnsku_count = fnsku_cells.count() if fnsku_cells else 0
            
            logger.info(
                f"检查仓库信息: 店铺列发现 {shop_count} 个单元格, FNSKU列发现 {fnsku_count} 个单元格"
            )
            
            if shop_count == 0 or fnsku_count == 0:
                logger.warning(
                    "未能定位到仓库信息单元格，请检查 colid 是否正确"
                )
                return True # 没找到也视为“需要处理”

            shop_texts = shop_cells.all_text_contents()
            fnsku_texts = fnsku_cells.all_text_contents()

            for i, text in enumerate(shop_texts):
                logger.info(f"店铺列第 {i+1} 行文本: {self._normalize_text(text)}")
            for i, text in enumerate(fnsku_texts):
                logger.info(f"FNSKU列第 {i+1} 行文本: {self._normalize_text(text)}")

            # Check if any text is empty or just whitespace
            for i, text in enumerate(shop_texts):
                if not text or not text.strip():
                    logger.info(f"第 {i+1} 行店铺信息为空")
                    return True
            for i, text in enumerate(fnsku_texts):
                if not text or not text.strip():
                    logger.info(f"第 {i+1} 行FNSKU信息为空")
                    return True
            
            logger.info("所有仓库信息已填入")
            return False
        except Exception as e:
            logger.error(f"检查仓库信息时发生异常: {e}")
        return True # Assume empty if we can't read it

    def wait_for_table_ready(self, timeout_ms: int = 10000) -> bool:
        """
        Wait until vxe table has rows or shows empty state.
        """
        selectors = []
        selectors.extend(LINGXING_SELECTORS.get("vxe_table_rows", []))
        selectors.extend(LINGXING_SELECTORS.get("vxe_table_empty", []))
        if not selectors:
            return True
        return wait_for_any_selector(self.page, selectors, timeout_ms=timeout_ms)

    def fix_empty_warehouse_info(self, order_no: str) -> None:
        """
        Iterate through warehouse cells, and if empty, click and select the first valid option.
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        
        # Define columns to check: (Cell Selector Key, Dropdown Option Selector Key, Column Name)
        columns = [
            ("warehouse_shop_cells", "warehouse_shop_dropdown_options", "发货仓库店铺"),
            ("warehouse_fnsku_cells", "warehouse_fnsku_dropdown_options", "发货仓库FNSKU")
        ]
        
        for cell_selector_key, opt_selector_key, col_name in columns:
            try:
                # Need to re-query elements every time because DOM might update
                cells = self.page.locator(LINGXING_SELECTORS.get(cell_selector_key)[0])
                count = cells.count()
                
                for i in range(count):
                    # Re-locate the specific cell
                    cell = cells.nth(i)
                    if not cell.is_visible():
                        continue
                        
                    text = cell.text_content()
                    if not text or not text.strip():
                        logger.info(f"{order_no}: 第 {i+1} 行 '{col_name}' 为空，尝试自动修复")
                        
                        # 1. Click the cell (or the input wrapper inside it)
                        wrapper = cell.locator(".el-input__inner").first
                        if wrapper.count() > 0:
                            wrapper.click()
                        else:
                            cell.click()
                        
                        human_pause(500, 800) # Give more time for dropdown
                        
                        # 2. Select corresponding valid option
                        if click_first_visible(self.page, LINGXING_SELECTORS.get(opt_selector_key, [])):
                             logger.info(f"{order_no}: 已选择 '{col_name}' 的默认选项")
                        else:
                             logger.warning(f"{order_no}: 未找到 '{col_name}' 可用的下拉选项")
                        
                        human_pause(300, 600)
            except Exception as e:
                logger.error(f"{order_no}: 修复 '{col_name}' 失败: {e}")

    def adjust_warehouse_logic(self, order_no: str) -> None:
        """
        Logic: Click adjust -> Check -> If empty, click adjust again -> Final check & Fix.
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        
        # 1. Click Adjust (First time) before checking
        if self.click_adjust():
            logger.info(f"{order_no}: 已点击调整按钮，等待数据加载...")
            wait_for_stability(self.page)
        else:
            logger.warning(f"{order_no}: 未能点击调整按钮，请核实页面状态")
            return

        logger.info(f"{order_no}: 开始检查并发货仓库信息...")

        # 2. Check and Click Adjust (Second time)
        if self.has_empty_warehouse_info():
            logger.info(f"{order_no}: 首次调整后检测到空值，尝试再次点击调整按钮...")
            self.click_adjust()
            wait_for_stability(self.page)
            
            # 3. Final Check & Fix
            if self.has_empty_warehouse_info():
                logger.warning(f"{order_no}: 经过两次调整仍有空值，启动手动补全逻辑...")
                self.fix_empty_warehouse_info(order_no)
            else:
                logger.info(f"{order_no}: 第二次调整后仓库信息已补全")
        else:
             logger.info(f"{order_no}: 首次调整后仓库信息已填入，无需进一步处理")

    def process_shipment(
        self,
        order_no: str,
        logistics_no: str,
        tracking_no: str,
        expected_transport_mode: Optional[str] = None,
        expected_logistics_channel: Optional[str] = None,
        expected_logistics_provider: Optional[str] = None,
        download_packing_template: bool = True,
        download_dir: Optional[str] = None,
    ) -> tuple[bool, str]:
        """
        Process a single shipment: navigate, edit, get packing, check quantity, fill logistics.
        
        Args:
            order_no: Shipment order number.
            logistics_no: Logistics provider number.
            tracking_no: Tracking number.
            expected_transport_mode: Expected transport mode (e.g., "海派").
            expected_logistics_channel: Expected logistics channel (e.g., "海派").
            expected_logistics_provider: Expected logistics provider name.
        
        Returns:
            Tuple of (success, message)
        """
        try:
            # 1. Navigate to shipment edit page directly
            self.navigate_to_shipment_edit(order_no)

            # 2.5 New Logic: Adjust warehouse info
            self.adjust_warehouse_logic(order_no)
            
            # 3. Click packing info tab
            # Ensure the edit page is ready before clicking
            if not wait_for_any_selector(self.page, LINGXING_SELECTORS.get("packing_info_tab", []), timeout_ms=10000):
                return (False, f"{order_no}: 发货单编辑页未就绪，找不到装箱信息选项卡")
            if not self.click_packing_info_tab():
                return (False, f"{order_no}: 无法点击装箱信息选项卡")
            wait_for_stability(self.page)
            
            # 4. Click get shipment packing
            if not self.click_get_shipment_packing():
                return (False, f"{order_no}: 无法点击获取货件装箱按钮")
            human_pause(1000, 2000)
            
            # 5. Confirm get shipment packing
            if not self.confirm_get_shipment_packing():
                return (False, f"{order_no}: 无法点击确定按钮")
            # Wait for success message before continuing
            if not wait_for_any_selector(
                self.page,
                LINGXING_SELECTORS.get("get_shipment_packing_success_msg", []),
                timeout_ms=15000,
            ):
                return (False, f"{order_no}: 未检测到获取货件装箱成功提示")
            wait_for_stability(self.page)

            if download_packing_template:
                ok, msg, template_path = self.download_packing_template(order_no, download_dir=download_dir)
                if not ok:
                    return (False, f"{order_no}: {msg}")
                # Click confirm button on edit page after download
                if not self.click_confirm_button():
                    return (False, f"{order_no}: 无法点击确定按钮")
                wait_for_stability(self.page)

                if not self.edit_packing_template(order_no, template_path):
                    return (False, f"{order_no}: 装箱清单模板编辑失败")

                # Re-enter edit page and upload template
                self.navigate_to_shipment_edit(order_no)
                if not wait_for_any_selector(self.page, LINGXING_SELECTORS.get("packing_info_tab", []), timeout_ms=10000):
                    return (False, f"{order_no}: 发货单编辑页未就绪，找不到装箱信息选项卡")
                if not self.click_packing_info_tab():
                    return (False, f"{order_no}: 无法点击装箱信息选项卡")
                wait_for_stability(self.page)
                if not self.upload_packing_template(template_path):
                    return (False, f"{order_no}: 装箱清单模板上传失败")
                # Wait until upload/download buttons disappear as success signal
                if not self.wait_packing_template_upload_complete():
                    return (False, f"{order_no}: 装箱清单模板上传后按钮未消失")
                wait_for_stability(self.page)
            
            is_match, msg = self.check_quantity_match()
            if not is_match:
                return (False, f"{order_no}: {msg}")
            
            # 7. Click logistics info tab
            if not self.click_logistics_info_tab():
                return (False, f"{order_no}: 无法点击物流信息选项卡")
            wait_for_stability(self.page)

            # 7.5 Check logistics profile before filling numbers (if expected values provided)
            if (
                expected_transport_mode is not None
                or expected_logistics_channel is not None
                or expected_logistics_provider is not None
            ):
                ok, msg = self.check_logistics_profile(
                    expected_transport_mode=expected_transport_mode,
                    expected_logistics_channel=expected_logistics_channel,
                    expected_logistics_provider=expected_logistics_provider,
                    auto_fix=True,
                )
                if not ok:
                    return (False, f"{order_no}: {msg}")
            
            # 8. Fill logistics numbers
            if not self.fill_logistics_numbers(logistics_no, tracking_no):
                return (False, f"{order_no}: 无法填写物流单号")
            
            # 9. Confirm (Temporarily disabled for testing)
            if not self.confirm_logistics_info():
                return (False, f"{order_no}: 无法点击确定保存")
            wait_for_stability(self.page)
            if not wait_for_any_selector(
                self.page,
                LINGXING_SELECTORS.get("edit_button", []),
                timeout_ms=10000,
            ):
                return (False, f"{order_no}: 保存后未检测到编辑按钮，可能未保存成功")
            
            return (True, f"{order_no}: 处理成功")
            
        except Exception as e:
            return (False, f"{order_no}: 处理异常 - {str(e)}")

    def download_packing_template(
        self,
        order_no: Optional[str] = None,
        *,
        download_dir: Optional[str] = None,
    ) -> tuple[bool, str, Optional[str]]:
        """
        Click '下载装箱清单模板' and save the Excel file if possible.
        """
        import os
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

        try:
            button_selectors = LINGXING_SELECTORS.get("packing_template_download_button", [])
            if not button_selectors:
                return (False, "未配置下载装箱清单模板按钮选择器", None)

            try:
                with self.page.expect_download(timeout=8000) as download_info:
                    if not click_first_visible(self.page, button_selectors):
                        return (False, "无法点击下载装箱清单模板按钮", None)
                download = download_info.value
                if download_dir is None:
                    download_dir = os.getcwd()
                os.makedirs(download_dir, exist_ok=True)
                filename = download.suggested_filename or f"{order_no or 'packing'}_装箱清单模板.xlsx"
                save_path = os.path.join(download_dir, filename)
                download.save_as(save_path)
                return (True, f"装箱清单模板已下载: {save_path}", save_path)
            except PlaywrightTimeoutError:
                if not click_first_visible(self.page, button_selectors):
                    return (False, "无法点击下载装箱清单模板按钮", None)
                human_pause(2000, 4000)
                return (False, "未捕获到下载文件，无法继续上传", None)
        except Exception as e:
            return (False, f"下载装箱清单模板失败: {str(e)}", None)

    def upload_packing_template(self, file_path: str) -> bool:
        """Upload packing template via file input."""
        if not file_path:
            return False
        try:
            file_input_selectors = LINGXING_SELECTORS.get("packing_template_upload_input", [])
            if not file_input_selectors:
                return False
            for selector in file_input_selectors:
                try:
                    locator = self.page.locator(selector)
                    if locator.count() > 0:
                        locator.first.set_input_files(file_path)
                        human_pause(1000, 2000)
                        return True
                except Exception:
                    continue
        except Exception:
            return False
        return False

    def wait_packing_template_upload_complete(self, timeout_ms: int = 15000) -> bool:
        """
        Wait until upload and download template buttons disappear.
        """
        import time

        start = time.monotonic()
        upload_btn = LINGXING_SELECTORS.get("packing_template_upload_input", [])
        download_btn = LINGXING_SELECTORS.get("packing_template_download_button", [])
        selectors = upload_btn + download_btn
        if not selectors:
            return True

        while (time.monotonic() - start) * 1000 < timeout_ms:
            all_gone = True
            for selector in selectors:
                try:
                    loc = self.page.locator(selector)
                    if loc.count() > 0 and loc.first.is_visible():
                        all_gone = False
                        break
                except Exception:
                    continue
            if all_gone:
                return True
            human_pause(300, 600)
        return False

    def click_confirm_button(self) -> bool:
        """Click generic confirm button on the edit page."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("confirm_button", []),
        )

    def edit_packing_template(self, order_no: str, template_path: Optional[str]) -> bool:
        """
        Edit packing template using LingXing API box_list data.
        """
        import logging
        import os
        from openpyxl import load_workbook

        logger = logging.getLogger("LingXingPortal")
        if not template_path or not os.path.exists(template_path):
            logger.error(f"{order_no}: 装箱清单模板文件不存在: {template_path}")
            return False

        try:
            from core.integrations.lingxing_client import LingXingClient
            from core.integrations import config as api_config
        except Exception as exc:
            logger.error(f"{order_no}: 导入领星API失败: {exc}")
            return False

        if not api_config.LINGXING_API_KEY or not api_config.LINGXING_TOKEN_URL:
            logger.error(f"{order_no}: 领星OpenAPI配置缺失，无法编辑装箱清单模板")
            return False

        client = LingXingClient(
            host=api_config.LINGXING_API_HOST,
            app_id=api_config.LINGXING_API_KEY,
            app_secret=api_config.LINGXING_API_SECRET,
            token_url=api_config.LINGXING_TOKEN_URL,
            token_key=api_config.LINGXING_TOKEN_REQUEST_KEY,
            ssl_verify=api_config.LINGXING_SSL_VERIFY,
        )

        try:
            data = self._run_async(client.get_shipment_data(order_no))
        except Exception as exc:
            logger.error(f"{order_no}: 获取领星发货单数据失败: {exc}")
            return False

        box_list = data.get("box_list") or []
        if not box_list:
            logger.error(f"{order_no}: box_list 为空，无法编辑装箱清单模板")
            return False

        try:
            wb = load_workbook(template_path)
            if "Sheet1" not in wb.sheetnames:
                logger.error(f"{order_no}: 模板缺少 Sheet1")
                return False
            ws = wb["Sheet1"]

            e1_value = ws.cell(row=1, column=5).value
            e1_value = str(e1_value).strip() if e1_value is not None else ""

            # 如果为“每箱一款SKU”，按固定列写入
            if e1_value == "每箱一款SKU":
                sku_col = 9  # I
                qty_col = 13  # M 单箱数量
                box_num_col = 14  # N 箱数
                weight_col = 15  # O 箱子毛重（kg）
                length_col = 16  # P 箱子长度（cm）
                width_col = 17  # Q 箱子宽度（cm）
                height_col = 18  # R 箱子高度（cm）

                def _to_int(value):
                    try:
                        return int(float(str(value)))
                    except Exception:
                        return 0

                # Build SKU->row mapping (header at row 3, data from row 4)
                sku_row_map = {}
                for row in range(4, ws.max_row + 1):
                    val = ws.cell(row=row, column=sku_col).value
                    if val is None:
                        continue
                    sku = str(val).strip()
                    if sku and sku not in sku_row_map:
                        sku_row_map[sku] = row

                for box in box_list:
                    box_num = box.get("box_num")
                    for sku_item in box.get("box_skus") or []:
                        sku = str(sku_item.get("sku") or "").strip()
                        qty = sku_item.get("quantity_in_case")
                        if not sku:
                            continue
                        row = sku_row_map.get(sku)
                        if row is None:
                            row = ws.max_row + 1
                            ws.cell(row=row, column=sku_col, value=sku)
                            sku_row_map[sku] = row

                        existing_qty = ws.cell(row=row, column=qty_col).value
                        if existing_qty in (None, ""):
                            ws.cell(row=row, column=qty_col, value=qty)

                        existing_box_num = ws.cell(row=row, column=box_num_col).value
                        ws.cell(row=row, column=box_num_col, value=_to_int(existing_box_num) + _to_int(box_num))

                        ws.cell(row=row, column=weight_col, value=box.get("cg_box_weight"))
                        ws.cell(row=row, column=length_col, value=box.get("cg_box_length"))
                        ws.cell(row=row, column=width_col, value=box.get("cg_box_width"))
                        ws.cell(row=row, column=height_col, value=box.get("cg_box_height"))

                wb.save(template_path)
                logger.info(f"{order_no}: 装箱清单模板已更新")
                return True

            # 默认逻辑：每箱多款SKU
            # Build SKU->row mapping (SKU column is I, header at row 6)
            sku_col = 9  # I
            sku_row_map = {}
            for row in range(7, ws.max_row + 1):
                val = ws.cell(row=row, column=sku_col).value
                if val is None:
                    continue
                sku = str(val).strip()
                if sku and sku not in sku_row_map:
                    sku_row_map[sku] = row

            start_col = 13  # M
            for idx, box in enumerate(box_list):
                col = start_col + idx
                ws.cell(row=1, column=col, value=box.get("cg_box_weight"))
                ws.cell(row=2, column=col, value=box.get("cg_box_length"))
                ws.cell(row=3, column=col, value=box.get("cg_box_width"))
                ws.cell(row=4, column=col, value=box.get("cg_box_height"))
                ws.cell(row=5, column=col, value=box.get("box_num"))

                for sku_item in box.get("box_skus") or []:
                    sku = str(sku_item.get("sku") or "").strip()
                    qty = sku_item.get("quantity_in_case")
                    if not sku:
                        continue
                    row = sku_row_map.get(sku)
                    if row is None:
                        logger.error(f"{order_no}: 模板中未找到 SKU: {sku}")
                        return False
                    ws.cell(row=row, column=col, value=qty)

            # Delete remaining columns after last box
            last_used_col = start_col + len(box_list) - 1
            max_col = ws.max_column
            if max_col > last_used_col:
                ws.delete_cols(last_used_col + 1, max_col - last_used_col)

            wb.save(template_path)
            logger.info(f"{order_no}: 装箱清单模板已更新")
            return True
        except Exception as exc:
            logger.error(f"{order_no}: 编辑装箱清单模板失败: {exc}")
            return False

    # ==================== 清关资料相关方法 ====================


    def navigate_to_customs_clear(self, order_no: str) -> None:
        """
        Navigate to customs clearance add page.

        Args:
            order_no: Shipment order number (e.g., SP260119003).
        """
        url = f"{config.LINGXING_CUSTOMS_CLEAR_URL}{order_no}"
        self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
        wait_for_stability(self.page)

    def click_customs_add_button(self) -> bool:
        """Click the add button to submit customs clearance form."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_add_button", []),
        )

    def _get_customs_receiver_section(self):
        for selector in LINGXING_SELECTORS.get("customs_receiver_section", []):
            try:
                locator = self.page.locator(selector)
                if locator.count() > 0:
                    return locator.first
            except Exception:
                continue
        return None

    def _read_input_value(self, locator) -> str:
        try:
            if locator is None or locator.count() == 0:
                return ""
            
            # 方法1: 尝试获取 input 的 value
            try:
                value = locator.first.input_value()
                value = self._normalize_text(value)
                if value:
                    return value
            except Exception:
                pass
            
            # 方法2: 尝试获取 attribute value
            value = self._normalize_text(locator.first.get_attribute("value"))
            if value:
                return value

            # 方法3 Fallback: 如果 input 为空，仅在其直接相关的容器（el-input）内查找 span.form-col-text
            # 这样可以避免误抓取到同一个 col-content 大容器内其他字段的 span（如仓库代码）
            try:
                # 尝试在 input 所在的直接 el-input 容器内寻找，如果不存在则说明该字段确实没有值
                el_input_parent = locator.first.locator("xpath=ancestor::div[contains(@class,'el-input')]")
                if el_input_parent.count() > 0:
                    span_node = el_input_parent.first.locator("xpath=..//span[contains(@class,'form-col-text')]")
                    if span_node.count() > 0:
                        return self._normalize_text(span_node.first.text_content())
            except Exception:
                pass

            return ""
        except Exception:
            return ""

    def _cell_has_image(self, cell) -> bool:
        try:
            if cell is None or cell.count() == 0:
                return False
            images = cell.first.locator("img")
            for idx in range(images.count()):
                src = images.nth(idx).get_attribute("src")
                if self._normalize_text(src):
                    return True
        except Exception:
            return False
        return False

    def _cell_has_value(self, cell) -> bool:
        try:
            if cell is None or cell.count() == 0:
                return False
            target = cell.first
            inputs = target.locator(".el-input__inner")
            if inputs.count() > 0:
                for idx in range(inputs.count()):
                    # 使用 input_value() 获取动态渲染的实时值
                    try:
                        val = self._normalize_text(inputs.nth(idx).input_value())
                    except Exception:
                        val = self._normalize_text(inputs.nth(idx).get_attribute("value"))
                    if not val:
                        label = target.locator(".fake-placeholder-hidden, .fake-select-label")
                        if label.count() > 0:
                            label_text = self._normalize_text(label.first.text_content())
                            if label_text:
                                continue
                        return False
                return True
            text = self._normalize_text(target.text_content())
            return bool(text)
        except Exception:
            return False

    def _cell_debug_text(self, cell, kind: str) -> str:
        """Best-effort read cell content for debugging missing fields."""
        try:
            if cell is None or cell.count() == 0:
                return ""
            target = cell.first
            if kind == "image":
                images = target.locator("img")
                if images.count() == 0:
                    return ""
                srcs = []
                for idx in range(images.count()):
                    src = self._normalize_text(images.nth(idx).get_attribute("src"))
                    if src:
                        srcs.append(src)
                return ", ".join(srcs)
            inputs = target.locator(".el-input__inner")
            if inputs.count() > 0:
                values = []
                for idx in range(inputs.count()):
                    # 优先使用 input_value() 获取实时值
                    try:
                        val = self._normalize_text(inputs.nth(idx).input_value())
                    except Exception:
                        val = self._normalize_text(inputs.nth(idx).get_attribute("value"))
                    if val:
                        values.append(val)
                if values:
                    return " / ".join(values)
                label = target.locator(".fake-placeholder-hidden, .fake-select-label")
                if label.count() > 0:
                    text = self._normalize_text(label.first.text_content())
                    if text:
                        return text
                return ""
            return self._normalize_text(target.text_content())
        except Exception:
            return ""

    def _find_select_by_placeholder(self, section, placeholder_text: str):
        if section is None:
            return None
        try:
            return section.locator(
                f'.//div[contains(@class,"el-select")][.//span[contains(@class,"fake-placeholder") and normalize-space(.)="{placeholder_text}"]]'
            )
        except Exception:
            return None

    def _read_select_value(self, select_locator) -> str:
        try:
            if select_locator is None or select_locator.count() == 0:
                return ""
            select = select_locator.first
            
            # 方法1: 尝试从 el-select 的仿真 label 中读取 (US, CA 等)
            label_selectors = [".fake-select-label", ".fake-placeholder-hidden", ".el-input__inner"]
            for s in label_selectors:
                label = select.locator(s)
                if label.count() > 0:
                    text = self._normalize_text(label.first.text_content())
                    # 排除掉占位符本身
                    if text and text not in ["国家", "省州"]:
                        return text
            
            # 方法2: 尝试属性
            input_el = select.locator(".el-input__inner")
            if input_el.count() > 0:
                value = self._normalize_text(input_el.first.get_attribute("value"))
                if value and value not in ["国家", "省州"]:
                    return value
        except Exception:
            return ""
        return ""

    def validate_customs_receiver_info(self, order_no: str) -> tuple[bool, str]:
        import logging
        import time
        logger = logging.getLogger("LingXingPortal")

        section = self._get_customs_receiver_section()
        if section is None:
            return (False, f"{order_no}: 未定位到收货仓库信息区域")
        if not wait_for_any_selector(self.page, LINGXING_SELECTORS.get("customs_receiver_loaded", []), timeout_ms=10000):
            return (False, f"{order_no}: 收货仓库信息未加载完成")

        name_val = country_val = province_val = city_val = zip_val = addr1_val = ""
        deadline = time.time() + 10
        
        # 调试：列出该区域内所有 input 的 placeholder 和 value
        try:
            all_inputs = section.locator("xpath=.//input").all()
            logger.info(f"{order_no}: 调试 - 发现 {len(all_inputs)} 个 input 元素")
            for idx, inp in enumerate(all_inputs):
                placeholder = inp.get_attribute("placeholder") or ""
                val = self._read_input_value(inp)
                logger.info(f"{order_no}: 调试 - Input[{idx}]: placeholder='{placeholder}', current_value='{val}'")
        except Exception as e:
            logger.warning(f"{order_no}: 调试日志收集失败: {e}")

        while True:
            # 使用更宽泛的选择器或直接遍历
            try:
                # 重新扫描所有 input 以获取最新值
                current_inputs = section.locator("xpath=.//input").all()
                empty_placeholder_inputs = []
                
                for inp in current_inputs:
                    p = (inp.get_attribute("placeholder") or "").strip()
                    v = self._read_input_value(inp)
                    if not v: continue
                    
                    if p == "名称": 
                        name_val = v
                    elif p == "城市": 
                        city_val = v
                    elif p == "邮编": 
                        zip_val = v
                    elif p == "详细地址1": 
                        addr1_val = v
                    elif not p:
                        # 记录没有 placeholder 的 input，通常按顺序是 国家、省州
                        empty_placeholder_inputs.append(v)
                
                # 如果正常提取逻辑没拿到 国家/省州，尝试从空 placeholder 列表中按顺序提取
                if not country_val and len(empty_placeholder_inputs) >= 1:
                    country_val = empty_placeholder_inputs[0]
                if not province_val and len(empty_placeholder_inputs) >= 2:
                    province_val = empty_placeholder_inputs[1]

                # Fallback: 尝试旧的选择框定位方式（如果上面没拿到）
                if not country_val:
                    country_val = self._read_select_value(self._find_select_by_placeholder(section, "国家"))
                if not province_val:
                    province_val = self._read_select_value(self._find_select_by_placeholder(section, "省州"))
                    
            except Exception as e:
                logger.warning(f"{order_no}: 循环提取异常: {e}")

            # 如果已经获取到关键信息，或者超时，则退出
            if all([name_val, country_val, province_val, city_val, zip_val, addr1_val]):
                break
            if time.time() >= deadline:
                break
            human_pause(300, 600)

        missing = []
        if not name_val:
            missing.append("名称")
        # 兼容性处理：如果读取到了 "国家" 或 "省州" 占位符，视为未填写
        if not country_val or country_val == "国家":
            missing.append("国家")
        if not province_val or province_val == "省州":
            missing.append("省州")
        if not city_val:
            missing.append("城市")
        if not zip_val:
            missing.append("邮编")
        if not addr1_val:
            missing.append("详细地址1")

        logger.info(
            f"{order_no}: 收货仓库信息读取: 名称={name_val}, 国家={country_val}, 省州={province_val}, 城市={city_val}, 邮编={zip_val}, 详细地址1={addr1_val}"
        )

        if missing:
            for field in missing:
                field_val = {
                    "名称": name_val,
                    "国家": country_val,
                    "省州": province_val,
                    "城市": city_val,
                    "邮编": zip_val,
                    "详细地址1": addr1_val,
                }.get(field, "")
                logger.info(f"{order_no}: 收货仓库信息缺失 {field}，当前值='{field_val}'")
            return (False, f"{order_no}: 收货仓库信息不完整: {', '.join(missing)}")

        return (True, f"{order_no}: 收货仓库信息完整")

    def _get_customs_product_header_map(self) -> dict[str, int]:
        """
        动态获取清关产品表格的表头位置映射。
        
        Returns:
            dict: {"列名": 1-based 索引}
        """
        import logging
        logger = logging.getLogger("LingXingPortal")
        header_map = {}
        try:
            header_row = self.page.locator('//table[contains(@class,"vxe-table--header")]//tr[contains(@class,"vxe-header--row")]/th')
            for idx in range(header_row.count()):
                th = header_row.nth(idx)
                text = self._normalize_text(th.text_content())
                if text:
                    # 处理多行表头的情况，取第一行关键词
                    key = text.split()[0] if ' ' in text else text
                    header_map[key] = idx + 1  # 1-based
            logger.info(f"动态表头识别: {header_map}")
        except Exception as e:
            logger.warning(f"动态表头识别失败: {e}")
        return header_map

    def validate_customs_product_table(self, order_no: str) -> tuple[bool, str]:
        import logging
        logger = logging.getLogger("LingXingPortal")

        row_selector = '//table[contains(@class,"vxe-table--body")]//tr[contains(@class,"vxe-body--row")]'
        if not wait_for_any_selector(self.page, [row_selector], timeout_ms=10000):
            return (False, f"{order_no}: 未找到清关产品信息表格")

        rows = self.page.locator(row_selector)
        if rows.count() == 0:
            return (False, f"{order_no}: 清关产品信息表格为空")

        # 动态获取表头映射
        header_map = self._get_customs_product_header_map()

        # 定义需要校验的字段（名称, 默认索引, 类型）
        # 索引会优先使用动态识别的结果，不在表头中的字段回退为默认值
        field_definitions = [
            ("货件编码", 13, "text"),
            ("内部编码Reference", 14, "value"),  # 表头是 "内部编码Reference Id"
            ("产品图片", 15, "image"),
            ("中文品名", 16, "value"),
            ("英文品名", 17, "value"),
            ("产品型号", 18, "text"),
            ("品牌", 19, "text"),
            ("中文材质", 20, "text"),
            ("英文材质", 21, "text"),
            ("中文用途", 22, "text"),
            ("英文用途", 23, "text"),
            ("FNSKU", 24, "value"),
            ("国外进口清关", 25, "text"),
        ]

        # 构建实际使用的字段列表，优先使用动态映射
        fields = []
        for field_name, default_idx, kind in field_definitions:
            # 尝试在 header_map 中找到匹配的列（部分匹配即可）
            col_idx = default_idx
            for hdr_text, hdr_idx in header_map.items():
                if field_name in hdr_text or hdr_text in field_name:
                    col_idx = hdr_idx
                    break
            fields.append((field_name, col_idx, kind))

        missing_rows = []
        for row_idx in range(rows.count()):
            row = rows.nth(row_idx)
            missing_fields = []
            row_values = []  # 收集当前行的所有字段值用于调试输出
            for field_name, col_index, kind in fields:
                cell = row.locator(f"xpath=./td[{col_index}]")
                ok = False
                debug_text = self._cell_debug_text(cell, kind)
                row_values.append(f"{field_name}='{debug_text}'")
                if kind == "image":
                    ok = self._cell_has_image(cell)
                else:
                    ok = self._cell_has_value(cell)
                if not ok:
                    missing_fields.append(field_name)
                    logger.info(
                        f"{order_no}: 清关产品信息缺失 第{row_idx + 1}行 {field_name}，当前值='{debug_text}'"
                    )
            # 输出当前行的所有字段值
            logger.info(f"{order_no}: 第{row_idx + 1}行数据: {', '.join(row_values)}")
            if missing_fields:
                missing_rows.append(f"第{row_idx + 1}行: {', '.join(missing_fields)}")

        if missing_rows:
            logger.warning(f"{order_no}: 清关产品信息不完整: {'; '.join(missing_rows)}")
            return (False, f"{order_no}: 清关产品信息不完整: {'; '.join(missing_rows)}")

        logger.info(f"{order_no}: 清关产品信息校验通过")
        return (True, f"{order_no}: 清关产品信息完整")


    def add_customs_clearance(self, order_no: str) -> tuple[bool, str]:
        """
        Complete customs clearance form for a shipment.

        Steps:
            1. Navigate to customs clearance page
            2. Click add button

        Args:
            order_no: Shipment order number (e.g., SP260119003).

        Returns:
            Tuple of (success: bool, message: str).
        """
        try:
            # 1. Navigate to customs clearance page
            self.navigate_to_customs_clear(order_no)
            human_pause(1000, 2000)

            # 1.1 Validate receiver warehouse info
            ok, msg = self.validate_customs_receiver_info(order_no)
            if not ok:
                return (False, msg)

            # 1.2 Validate product info table
            ok, msg = self.validate_customs_product_table(order_no)
            if not ok:
                return (False, msg)

            # 2. Click add button
            if not self.click_customs_add_button():
                return (False, f"{order_no}: 无法点击添加按钮")
            
            # 等待跳转到清关管理页面并加载出数据
            try:
                # 显式等待 URL 变更
                self.page.wait_for_url(config.LINGXING_CUSTOMS_MANAGE_URL, timeout=10000)
                # 等待表格第一行数据加载出来
                if not wait_for_any_selector(self.page, LINGXING_SELECTORS.get("customs_manage_table_row", []), timeout_ms=10000):
                    return (False, f"{order_no}: 添加后未在管理页面检测到数据行")
            except Exception as e:
                return (False, f"{order_no}: 添加后等待跳转管理页面超时或异常 - {str(e)}")

            human_pause(1000, 2000)
            return (True, f"{order_no}: 清关资料添加成功")

        except Exception as e:
            return (False, f"{order_no}: 添加清关资料异常 - {str(e)}")

    # ==================== 清关管理下载相关方法 ====================

    def navigate_to_customs_manage(self) -> None:
        """Navigate to customs clearance management page."""
        self.page.goto(config.LINGXING_CUSTOMS_MANAGE_URL, wait_until="domcontentloaded", timeout=30000)
        wait_for_stability(self.page)

    def click_filter_select(self) -> bool:
        """Click the filter condition dropdown."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_filter_select", []),
        )

    def select_related_order_filter(self) -> bool:
        """Select '关联单号' filter option."""
        human_pause(500, 1000)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_filter_option_related", []),
        )

    def fill_search_input(self, order_no: str) -> bool:
        """Fill the search input with order number."""
        human_pause(300, 600)
        return fill_first(
            self.page,
            LINGXING_SELECTORS.get("customs_search_input", []),
            order_no,
        )

    def click_search_button(self) -> bool:
        """Click the search button."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_search_button", []),
        )

    def click_select_all(self) -> bool:
        """Click the select all checkbox."""
        human_pause(500, 1000)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_select_all", []),
        )

    def click_download_button(self) -> bool:
        """Click the download button (批量下载)."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_download_button", []),
        )

    def click_download_template_select(self) -> bool:
        """Click the print template select dropdown in batch download dialog."""
        human_pause(300, 600)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_download_template_select", []),
        )

    def select_download_template_pingyi(self) -> bool:
        """Select '平谊国际下单模板' in batch download dialog."""
        human_pause(500, 1000)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_download_template_option_pingyi", []),
        )

    def click_download_confirm(self) -> bool:
        """Click the confirm download button in dialog."""
        human_pause(500, 1000)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_download_confirm", []),
        )

    def click_download_close(self) -> bool:
        """Click the close button after download."""
        human_pause(1000, 2000)
        return click_first_visible(
            self.page,
            LINGXING_SELECTORS.get("customs_download_close", []),
        )

    def download_customs_files(self, order_no: str, download_dir: Optional[str] = None) -> tuple[bool, str]:
        """
        Download customs clearance files for a shipment.

        Steps:
            1. Fill search input with order number
            2. Click search button
            3. Click select all
            4. Click download button
            5. Select print template (平谊国际下单模板)
            6. Click confirm download
            7. Click close button

        Args:
            order_no: Shipment order number (e.g., SP260119003).
            download_dir: Optional directory to save downloaded file.

        Returns:
            Tuple of (success: bool, message: str).
        """
        try:
            # 1. Fill search input
            if not self.fill_search_input(order_no):
                return (False, f"{order_no}: 无法填写搜索框")
            human_pause(300, 600)

            # 2. Click search button
            if not self.click_search_button():
                return (False, f"{order_no}: 无法点击查询按钮")
            wait_for_stability(self.page)
            
            # 等待表格数据加载完成（等待表格行出现）
            table_row_selector = '//table[contains(@class,"vxe-table--body")]//tr[contains(@class,"vxe-body--row")]'
            if not wait_for_any_selector(self.page, [table_row_selector], timeout_ms=15000):
                return (False, f"{order_no}: 搜索后未找到数据行，可能没有匹配的清关资料")
            human_pause(1000, 2000)

            # 3. Click select all
            if not self.click_select_all():
                return (False, f"{order_no}: 无法点击全选框")
            human_pause(500, 1000)


            # 4. Click download button
            if not self.click_download_button():
                return (False, f"{order_no}: 无法点击下载按钮")
            human_pause(500, 1000)

            # 5. Select print template
            if not self.click_download_template_select():
                return (False, f"{order_no}: 无法点击打印模板选择框")
            if not self.select_download_template_pingyi():
                return (False, f"{order_no}: 无法选择平谊国际下单模板")
            human_pause(500, 1000)

            # 6. Click confirm download (capture download if directory provided)
            if download_dir:
                import os
                os.makedirs(download_dir, exist_ok=True)
                try:
                    with self.page.expect_download(timeout=15000) as download_info:
                        if not self.click_download_confirm():
                            return (False, f"{order_no}: 无法点击确认下载按钮")
                    download = download_info.value
                    filename = download.suggested_filename or f"{order_no}_清关文件.xlsx"
                    save_path = os.path.join(download_dir, filename)
                    download.save_as(save_path)
                except Exception as e:
                    return (False, f"{order_no}: 下载清关文件异常 - {str(e)}")
            else:
                if not self.click_download_confirm():
                    return (False, f"{order_no}: 无法点击确认下载按钮")
                human_pause(2000, 4000)  # Wait for download to complete

            # 7. Click close button
            if not self.click_download_close():
                return (False, f"{order_no}: 无法点击关闭按钮")
            wait_for_stability(self.page)

            return (True, f"{order_no}: 清关文件下载成功")

        except Exception as e:
            return (False, f"{order_no}: 下载清关文件异常 - {str(e)}")

    def download_customs_files_batch(
        self,
        order_nos: list[str],
        download_dir: Optional[str] = None,
    ) -> list[tuple[bool, str]]:
        """
        Download customs clearance files for multiple shipments.

        Args:
            order_nos: List of shipment order numbers.

        Returns:
            List of (success, message) tuples for each order.
        """
        results = []
        
        # Navigate to customs manage page first
        self.navigate_to_customs_manage()
        human_pause(1000, 2000)
        
        # Set filter to '关联单号' once
        if not self.click_filter_select():
            return [(False, "无法点击筛选条件框")]
        human_pause(500, 1000)
        
        if not self.select_related_order_filter():
            return [(False, "无法选择关联单号筛选")]
        wait_for_stability(self.page)
        
        # Process each order
        for order_no in order_nos:
            result = self.download_customs_files(order_no, download_dir=download_dir)
            results.append(result)
            human_pause(500, 1000)
        
        return results
