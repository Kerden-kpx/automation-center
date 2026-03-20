# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""Excel registry and folder management utilities.

This module provides reusable utilities for:
1. Reading/writing entries in an Excel registry file
2. Creating folders on shared drives with standardized naming
3. Auto-incrementing code generation
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Optional, Tuple

from openpyxl import load_workbook


@dataclass
class RegistryConfig:
    """Configuration for an Excel registry."""

    excel_path: str
    """Path to the Excel registry file."""

    sheet_name: str
    """Name of the worksheet to use."""

    code_col_index: int
    """1-indexed column number for the code (e.g., 11 for column K)."""

    remark_col_index: int
    """1-indexed column number for the remark (e.g., 12 for column L)."""

    folder_base_dir: str
    """Base directory for creating folders."""

    default_remark: str = ""
    """Default remark value if not provided."""


def _find_last_value(ws, col_index: int) -> Tuple[Optional[int], Optional[str]]:
    """Find the last non-empty value in a column.

    Args:
        ws: openpyxl worksheet
        col_index: 1-indexed column number

    Returns:
        Tuple of (row_number, value) or (None, None) if not found
    """
    for row in range(ws.max_row, 0, -1):
        value = ws.cell(row=row, column=col_index).value
        if value is None:
            continue
        value_str = str(value).strip()
        if value_str:
            return row, value_str
    return None, None


def _increment_code(last_code: str) -> str:
    """Increment a code by finding trailing digits and adding 1.

    Examples:
        AGL2026-001 -> AGL2026-002
        INV-0099 -> INV-0100

    Args:
        last_code: The previous code string

    Returns:
        Incremented code with same format

    Raises:
        ValueError: If no trailing digits found
    """
    match = re.search(r"(\d+)$", last_code)
    if not match:
        raise ValueError(f"无法在编号末尾找到数字: {last_code!r}")
    number_str = match.group(1)
    prefix = last_code[:-len(number_str)]
    next_number = int(number_str) + 1
    return f"{prefix}{str(next_number).zfill(len(number_str))}"


def _get_sheet(workbook, sheet_name: str):
    """Get a worksheet by name, raising ValueError if not found."""
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    raise ValueError(f"未找到工作表: {sheet_name}")


def get_latest_entry(config: RegistryConfig) -> Tuple[str, str]:
    """Get the latest entry (code and remark) from the registry.

    Args:
        config: Registry configuration

    Returns:
        Tuple of (code, remark)

    Raises:
        ValueError: If no entries found in the code column
    """
    wb = load_workbook(config.excel_path, data_only=True)
    ws = _get_sheet(wb, config.sheet_name)

    last_row, last_code = _find_last_value(ws, config.code_col_index)
    if last_row is None or last_code is None:
        raise ValueError(f"列 {config.code_col_index} 未找到任何编号，无法获取最新记录。")

    remark_value = ws.cell(row=last_row, column=config.remark_col_index).value
    remark = str(remark_value).strip() if remark_value else config.default_remark
    return last_code, remark


def ensure_target_folder(
    code: str,
    remark: str,
    base_dir: str,
    create: bool = True,
) -> str:
    """Ensure a folder exists with standardized naming.

    Folder name format: "{code} {remark}"

    Args:
        code: The registry code
        remark: The remark/description
        base_dir: Base directory path
        create: Whether to create the folder if it doesn't exist

    Returns:
        Full path to the folder
    """
    folder_name = f"{code} {remark}".strip()
    target = os.path.join(base_dir, folder_name)
    if create:
        os.makedirs(target, exist_ok=True)
    return target


def create_next_entry(
    config: RegistryConfig,
    remark: Optional[str] = None,
) -> Tuple[str, str]:
    """Create a new entry in the registry and corresponding folder.

    This function:
    1. Reads the last code from the Excel file
    2. Generates the next code by incrementing
    3. Writes the new code and remark to the next row
    4. Creates a folder with the naming convention "{code} {remark}"

    Args:
        config: Registry configuration
        remark: Optional remark override (uses config.default_remark if None)

    Returns:
        Tuple of (new_code, folder_path)

    Raises:
        ValueError: If no entries found to increment from
    """
    remark = remark or config.default_remark

    wb = load_workbook(config.excel_path)
    ws = _get_sheet(wb, config.sheet_name)

    last_row, last_code = _find_last_value(ws, config.code_col_index)
    if last_row is None or last_code is None:
        raise ValueError(f"列 {config.code_col_index} 未找到任何编号，无法生成新编号。")

    next_code = _increment_code(last_code)
    new_row = last_row + 1
    ws.cell(row=new_row, column=config.code_col_index, value=next_code)
    ws.cell(row=new_row, column=config.remark_col_index, value=remark)
    wb.save(config.excel_path)

    folder_path = ensure_target_folder(next_code, remark, base_dir=config.folder_base_dir, create=True)
    return next_code, folder_path


def get_entry_folder_path(
    code: str,
    remark: str,
    base_dir: str,
) -> str:
    """Get the folder path for an existing entry without creating it.

    Args:
        code: The registry code
        remark: The remark/description
        base_dir: Base directory path

    Returns:
        Full path to the folder (may or may not exist)
    """
    return ensure_target_folder(code, remark, base_dir, create=False)
