# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""Invoice registry helpers for logistics."""

from __future__ import annotations

import re
from typing import Optional, Tuple

from openpyxl import load_workbook


def _find_last_value(ws, col_index: int) -> Tuple[Optional[int], Optional[str]]:
    for row in range(ws.max_row, 0, -1):
        value = ws.cell(row=row, column=col_index).value
        if value is None:
            continue
        value_str = str(value).strip()
        if value_str:
            return row, value_str
    return None, None


def _increment_code(last_code: str) -> str:
    match = re.search(r"(\d+)$", last_code)
    if not match:
        raise ValueError(f"无法在编号末尾找到数字: {last_code!r}")
    number_str = match.group(1)
    prefix = last_code[:-len(number_str)]
    next_number = int(number_str) + 1
    return f"{prefix}{str(next_number).zfill(len(number_str))}"


def _pick_sheet(workbook, sheet_name: Optional[str] = None):
    if sheet_name:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        raise ValueError(f"未找到工作表: {sheet_name}")
    return workbook.active


def register_invoice(
    shop: str,
    country: str,
    transport: str,
    excel_path: str,
    sheet_name: Optional[str] = None,
) -> dict:
    """Register a new invoice number for a shop and transport.

    Args:
        shop: 店铺名称（用于选择列）
        country: 国家
        transport: 运输方式（海派/海卡等）
        excel_path: 登记表路径
        sheet_name: 可选工作表名（默认活动表）

    Returns:
        dict with keys: code, remark, row
    """
    if not shop:
        raise ValueError("缺少店铺名称。")
    if not country:
        raise ValueError("缺少国家。")
    if not transport:
        raise ValueError("缺少运输方式。")

    shop_upper = shop.strip().upper()
    if shop_upper.startswith("EZARC"):
        code_col = 1  # A
        remark_col = 2  # B
    elif shop_upper.startswith("TOLESA"):
        code_col = 6  # F
        remark_col = 7  # G
    else:
        raise ValueError(f"未识别的店铺前缀: {shop}")

    transport_text = "Agl AMP" if transport.strip() == "海卡" else transport.strip()
    remark = f"{country} + {transport_text}"

    wb = load_workbook(excel_path)
    ws = _pick_sheet(wb, sheet_name)
    last_row, last_code = _find_last_value(ws, code_col)
    if last_row is None or not last_code:
        raise ValueError("登记表未找到历史编号，无法递增。")
    next_code = _increment_code(last_code)
    new_row = last_row + 1
    ws.cell(row=new_row, column=code_col, value=next_code)
    ws.cell(row=new_row, column=remark_col, value=remark)
    wb.save(excel_path)
    return {"code": next_code, "remark": remark, "row": new_row}
